#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
吉木萨尔芦草沟组 Rock-Eval 数据入库（Paper XLIII 中国湖相第六体系，2026-08-09）

数据源：Mendeley Data, Ma, Kuiyou; Hong, Pang; Junqing, Chen; Lin, Jiang (2024),
  "Data for Organic-Inorganic Lithofacies Subdivision Research in the Lucaogou Shale",
  V1, doi: 10.17632/sy6znr66dc.1（CC BY 4.0）
  Dataset.xlsx：118+ 样品、6 口井（W1-W6）、吉木萨尔凹陷芦草沟组。

处理：
  - 'Original Dataset' 表提取 Sample Code / Lithology / S1 / S2 / Tmax / HI / TOC / OSI
  - 'Cluster Result' 表提取 Mean buried depth (m)
  - 输出标准 CSV（列与 rockeval_chang7 等一致：Sample_ID,Depth_m,TOC_wt,S1_mgg,S2_mgg,Tmax_C）
  - 样本代码含井名（W1-W6），保留为体系内井分组字段
"""
import csv
import os
import re
import openpyxl
import numpy as np

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "data", "rockeval_jimsar_lucaogou", "Dataset.xlsx")
DST = os.path.join(BASE, "data", "rockeval_jimsar_lucaogou", "lucaogou_rockeval.csv")

def _tof(x):
    try:
        return float(x)
    except (ValueError, TypeError):
        return np.nan

def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    orig = list(wb["Original Dataset"].iter_rows(values_only=True))
    clus = list(wb["Cluster Result"].iter_rows(values_only=True))
    wb.close()

    # 深度索引：Cluster Result 第 5 列 Mean buried depth（m），第 3 列 Sample Code
    depth_map = {}
    for r in clus[1:]:
        if r[2] is None:
            continue
        code = str(r[2]).strip()
        dep = _tof(r[4])
        if np.isfinite(dep):
            depth_map[code] = dep

    # 列索引（Original Dataset）：0 Number,1 Sample Code,2 Lithology,22 S1,23 S2,24 Tmax,26 HI,27 TOC,28 OSI
    hdr = orig[0]
    idx = {h: i for i, h in enumerate(hdr)}
    out = []
    n_ok = 0
    well_stat = {}
    for r in orig[1:]:
        code = str(r[1]).strip() if r[1] is not None else ""
        if not code:
            continue
        s1, s2, tmax, toc = _tof(r[idx["S1(mg/g)"]]), _tof(r[idx["S2(mg/g)"]]), \
                            _tof(r[idx["Tmax(℃)"]]), _tof(r[idx["TOC(%)"]])
        if not (np.isfinite(toc) and np.isfinite(s1) and 0 < toc < 30 and s1 >= 0):
            continue
        well = "W" + code[1] if re.match(r"W\d", code) else ""
        dep = depth_map.get(code, np.nan)
        out.append([code, dep, toc, s1, s2, tmax, well, str(r[2]) if r[2] else ""])
        well_stat[well] = well_stat.get(well, 0) + 1
        n_ok += 1

    with open(DST, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Sample_ID", "Depth_m", "TOC_wt", "S1_mgg", "S2_mgg", "Tmax_C", "Well", "Lithology"])
        w.writerows(out)

    print("芦草沟组入库完成：%d 样品 -> %s" % (n_ok, DST))
    print("分井样品数：%s" % well_stat)
    print("深度覆盖：%.0f–%.0f m" % (np.nanmin([o[1] for o in out]), np.nanmax([o[1] for o in out])))
    toc = np.array([o[2] for o in out]); s1 = np.array([o[3] for o in out]); tm = np.array([o[5] for o in out])
    print("TOC %.2f–%.2f（中位 %.2f）| Tmax %.0f–%.0f（中位 %.0f）| OSI 中位 %.1f"
          % (toc.min(), toc.max(), np.median(toc), tm.min(), tm.max(), np.median(tm),
             np.median(s1 / toc * 100)))

if __name__ == "__main__":
    main()
